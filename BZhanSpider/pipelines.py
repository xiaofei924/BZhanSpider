# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
"""
用来对items里面提取的数据做进一步处理，如保存等
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment
from scrapy.exporters import CsvItemExporter

from BZhanSpider.replybean import replybean


class BzhanspiderPipeline(object):
    # 网友评论列
    video_comment_column = 12
    # up主回复列
    up_reply_column = 13
    # 网友互动列
    net_friend_reply_column = 14
    # 从第几行开始插入数据
    data_start_row = 3
    """列宽"""
    column_width = 20
    """行高"""
    row_height = 10

    """标题， 水平居中，垂直居中，不自动换行"""
    title_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
    """水平居中，垂直居中"""
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    """水平左对齐，垂直居中"""
    left_horizontal_center_vertical_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    """水平左对齐，垂直左对齐"""
    left_horizontal_and_vertical_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    def __init__(self):
        pass
        # self.ws.append(top_title)

    def open_spider(self, spider):
        if spider.name == 'BZhanSearchInfo':
            # 创建excel，填写表头
            self.wb = Workbook()
            self.ws = self.wb.active
            title = ['url', '标题']
            self.ws.append(title)
        if spider.name == 'BZhan':
            # pass
            # 初始化一个文件
            self.file_name = open("BZhanInfo.json", "w", encoding='utf-8')

            # 创建excel，填写表头
            self.wb = Workbook()
            self.ws = self.wb.active
            top_title = ['标题',
                         '栏目（舞蹈、鬼畜、科技等）',
                         '发表时间',
                         '排名',
                         '播放量',
                         '弹幕数',
                         '点赞数',
                         '投币数',
                         '收藏数',
                         '文字介绍',
                         '关键词tag',
                         '',
                         'UP主与网友互动内容',
                         '',
                         'UP主',
                         'UP主简介',
                         '认证分类',
                         '关注数',
                         '粉丝数',
                         '播放数']
            sub_title = ['',
                         '',
                         '',
                         '',
                         '',
                         '',
                         '',
                         '',
                         '',
                         '',
                         '',
                         '网友评论',
                         'UP主回复',
                         '网友互动',
                         '',
                         '',
                         '',
                         '',
                         '',
                         '']
            # self.ws.append(sub_title)

            title_len = len(top_title)

            """插入标题"""
            self.ws.cell(row=1, column=self.video_comment_column).value = 'UP主与网友互动内容'
            for i in range(1, title_len + 1):
                """标题水平居中"""
                self.ws.cell(row=1, column=i).alignment = self.center_alignment
                '''设置列宽'''
                self.ws.column_dimensions[chr(65 + i - 1)].width = self.column_width
                if i == 12 or i == 13 or i == 14:
                    self.ws.merge_cells(start_row=1, end_row=1, start_column=self.video_comment_column,
                                        end_column=self.net_friend_reply_column)

                    continue
                self.ws.merge_cells(start_row=1, end_row=2, start_column=i, end_column=i)
                self.ws.cell(row=1, column=i).value = top_title[i - 1]

            self.ws.cell(row=2, column=self.video_comment_column).value = '网友评论'
            self.ws.cell(row=2, column=self.up_reply_column).value = 'UP主回复'
            self.ws.cell(row=2, column=self.net_friend_reply_column).value = '网友互动'
            """标题水平居中"""
            self.ws.cell(row=2, column=self.video_comment_column).alignment = self.center_alignment
            self.ws.cell(row=2, column=self.up_reply_column).alignment = self.center_alignment
            self.ws.cell(row=2, column=self.net_friend_reply_column).alignment = self.center_alignment

    def process_item(self, item, spider):
        if spider.name == 'BZhanSearchInfo':
            search_item = dict(item)
            line = []
            file_name = ''
            if 'vlog_map' in search_item:
                vlog_map = search_item['vlog_map']
                for url, title in vlog_map.items():
                    line = [str(url), str(title)]
                    self.ws.append(line)
                file_name = 'BZhanSearchInfo_' + spider.keywords + '_去重版.xlsx'
            else:
                line = [item['url'], item['title']]
                file_name = 'BZhanSearchInfo_' + spider.keywords + '_全部版.xlsx'
            self.ws.append(line)
            self.wb.save(file_name)
            print('--------------------- BzhanspiderPipeline, process_item item end ----------------------\n')
            return item
        if spider.name == 'BZhan':
            print('--------------------- BzhanspiderPipeline, process_item  item start ----------------------\n')
            # print(item)
            if item is None:
                return
            item = dict(item)
            is_commet_or_reply = False

            if 'video_reply_map' in item:
                video_reply_map = item['video_reply_map']
                print('video_reply_map： \n')
                video_reply_map_len = len(video_reply_map)
                row = self.data_start_row  # 从数据开始行开始插入
                for key in video_reply_map:
                    replybean = video_reply_map[key]
                    # print(replybean.to_string())
                    cell = self.ws.cell(row=row, column=self.video_comment_column)
                    cell.alignment = self.center_alignment
                    # cell.value = 'rpid: ' + replybean.get_rpid_str() \
                    #              + ', floor: ' + replybean.get_floor() \
                    #              + ', content: ' + replybean.get_content()
                    cell.value = replybean.get_content()
                    row += 1
                is_commet_or_reply = True

            if 'up_reply_map' in item:
                up_reply_map = item['up_reply_map']
                # print('up_reply_map： \n')
                row = self.data_start_row  # 从数据开始行开始插入
                for key in up_reply_map:
                    replybean = up_reply_map[key]
                    # print(replybean.to_string())
                    cell = self.ws.cell(row=row, column=self.up_reply_column)
                    cell.alignment = self.center_alignment
                    # cell.value = 'rpid: ' + replybean.get_rpid_str() \
                    #              + ', root id: ' + replybean.get_root_str()\
                    #              + ', parent id: ' + replybean.get_parent_str()\
                    #              + ', floor: ' + replybean.get_floor()\
                    #              + ', content: ' + replybean.get_content()
                    cell.value = replybean.get_content()
                    row += 1
                is_commet_or_reply = True

            if 'net_friend_reply_map' in item:
                net_friend_reply_map = item['net_friend_reply_map']
                # print('net_friend_reply_map： \n')
                row = self.data_start_row  # 从数据开始行开始插入
                for key in net_friend_reply_map:
                    replybean = net_friend_reply_map[key]
                    # print(replybean.to_string())
                    cell = self.ws.cell(row=row, column=self.net_friend_reply_column)
                    cell.alignment = self.center_alignment
                    # cell.value = 'rpid: ' + replybean.get_rpid_str() \
                    #              + ', root id: ' + replybean.get_root_str()\
                    #              + ', parent id: ' + replybean.get_parent_str()\
                    #              + ', floor: ' + replybean.get_floor()\
                    #              + ', content: ' + replybean.get_content()
                    cell.value = replybean.get_content()
                    row += 1
                is_commet_or_reply = True

            if not is_commet_or_reply:
                # 这里是将item先转换成字典，在又字典转换成字符串
                # json.dumps转换时对中文默认使用的ascii编码.想输出真正的中文需要指定 ensure_ascii=False
                # 将最后的item 写入到文件中
                text = json.dumps(dict(item), ensure_ascii=False) + "\n"
                self.file_name.write(text)
                line = [item['title'],
                        item['column'],
                        item['publish_date'],
                        item['ranking'],
                        item['play_count'],
                        item['danmu_count'],
                        item['like_count'],
                        item['corn_count'],
                        item['favorite_count'],
                        item['text_introduce'],
                        item['keywords_tag'],
                        '',
                        '',
                        '',
                        item['up_name'],
                        item['up_introduction'],
                        item['up_certification'],
                        item['up_focus_count'],
                        item['up_fans_count'],
                        item['up_play_count']]
                self.ws.append(line)
                for i in range(1, len(line) + 1):
                    self.ws.cell(row=3, column=i).alignment = self.center_alignment

            # if spider.is_save:
            self.wb.save('BZhanVideoInfo_' + spider.oid + '.xlsx')
            print('--------------------- BzhanspiderPipeline, process_item item end ----------------------\n')
            return item

    def close_spider(self, spider):
        pass
        # self.exporter.finish_exporting()
        # self.file.close()
#
# def process_item(self, item, spider):
#     self.exporter.export_item(item)
#     return item
